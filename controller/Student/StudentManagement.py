import json
import re
from datetime import datetime

import pandas
import pytz
from flask import (
    Blueprint,
    # Blueprint is a way to organize a group of related views and other code
    # There will be 2 blueprints: one for authentication and one for posts function
    request,
    jsonify
)
from openpyxl import load_workbook

from controller.Authentication.Authentication import token_required
from db.oasis_entites import Student, User, Semester, Course

student = Blueprint('StudentManagement', __name__, url_prefix='/student')


@student.route('/create-record', methods=['POST'])
@token_required
def create(e):
    try:
        new_student = request.get_json()
        code = new_student.get('new_code')
        username = new_student.get('new_username')
        name = new_student.get('new_name')
        email = new_student.get('new_email')
        dob = new_student.get('new_dob')
        class_course = new_student.get('new_class_cource')
        new_course_id = new_student.get('new_course_id')
        permission = 'Sinh viên'
        create_at = datetime.now(pytz.timezone('Asia/Ho_Chi_Minh')).strftime("%Y-%m-%d %H:%M:%S")
        actived = new_student.get('new_actived')
        is_lock = new_student.get('new_is_lock')

        print(request.get_json())
        isStudentCreated = User.createRecord(str(username).strip(), str(name).strip(), str(email).strip(), create_at,
                                             permission, actived, is_lock, str(code).strip(), dob,
                                             str(class_course).strip(), new_course_id)
        if isStudentCreated is True:
            return jsonify({'status': 'success'}), 200
        else:
            return jsonify({'status': 'already-exist'}), 202
    except Exception as e:
        return jsonify({'status': 'bad-request', 'error_message': e.__str__()}), 400


@student.route('/import-excel', methods=['POST'])
@token_required
def import_excel(e):
    excel_file = request.files['student_list_excel']
    new_students = []
    update_students = []
    error_students = []
    course_id = request.form.get('course_id')
    try:
        if excel_file.content_type == 'application/vnd.ms-excel':  # xls file type
            # load data
            data = pandas.read_excel(request.files['student_list_excel'], encoding='ISO-8859-1')

            processed_data = json.dumps(data, ensure_ascii=False)

            excel_json = json.loads(processed_data)
            for item in excel_json['DSLMH']:
                print(item)
                print('\n')
            return jsonify({'status': 'success'}, excel_json), 200
        elif excel_file.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':  # xlsx file type
            data = load_workbook(excel_file)

            sheet = data.active
            # get max row count
            max_row = sheet.max_row + 1
            # get max column count
            max_column = sheet.max_column + 1
            print('BEBE')
            for i in range(12, max_row):
                student = {
                    'STT': None,
                    'Mã SV': None,
                    'Họ và tên': None,
                    'Ngày sinh': None,
                    'Lớp khóa học': None,
                    'Ghi chú': None,
                }
                # Iterate over the dict and all the columns
                for j, index in zip(range(1, max_column), student):
                    # add data to excel_data dict
                    student[index] = sheet.cell(row=i, column=j).value
                print(student, flush=True)
                # add students to database
                # check validation
                if student['STT'] is not None and student['Mã SV'] is not None:
                    check_code = re.search('^\d{8}$', str(student['Mã SV']).replace(' ', ''))
                    check_name = re.search("^[a-zA-Z_ÀÁÂÃÈÉÊÌÍÒÓÔÕÙÚĂĐĨŨƠàáâãèéêìíòóôõùúăđĩũơƯĂẠẢẤẦẨẪẬẮẰẲẴẶ" +
                                           "ẸẺẼỀẾỂưăạảấầẩẫậắằẳẵặẹẻẽềếểỄỆỈỊỌỎỐỒỔỖỘỚỜỞỠỢỤỦỨỪễệỉịọỏốồổỗộớờởỡợ" +
                                           "ụủứừỬỮỰỲỴÝỶỸửữựỳýỵỷỹ\s ]+$", str(student['Họ và tên']))
                    check_dob = re.search('^(((0)[1-9])|((1)[0-2]))(\/)([0-2][0-9]|(3)[0-1])(\/)\d{4}',
                                          str(datetime.strptime(student['Ngày sinh'], '%d/%m/%Y').strftime('%m/%d/%Y')))
                    # gender = re.search('(Nam|Nữ)', str(excel_data['gender']))
                    # courseID = re.search('^[K|k][1-9][0-9][A-Za-z]+[1-9]*', str(excel_data['courseID']).replace(' ', ''))
                    # subjectID = re.search('(^(([A-Z]|[a-z]){3})([1-9][(0-9)]{3})$)',
                    #                       str(excel_data['subjectID']).replace(' ', ''))
                    # status = re.search('(đủ điều kiện|không đủ điều kiện)', excel_data['status'].lower())
                    print(check_code, flush=True)
                    print(check_name, flush=True)
                    print(check_dob, flush=True)
                    # print(gender, flush=True)
                    # print(courseID, flush=True)
                    # print(subjectID, flush=True)
                    # print(status, flush=True)

                    if (check_code is not None) and (check_name is not None) and (check_dob is not None):
                        isStudentCreated = User.createRecord(str(student['Mã SV']).strip(),
                                                             str(student['Họ và tên']).strip(),
                                                             str(student['Mã SV']) + '@vnu.edu.vn',
                                                             datetime.now(pytz.timezone('Asia/Ho_Chi_Minh')).strftime(
                                                                 "%Y-%m-%d %H:%M:%S"),
                                                             'Sinh viên', 1, 0, str(student['Mã SV']).strip(),
                                                             datetime.strptime(student['Ngày sinh'], '%d/%m/%Y').strftime("%Y-%m-%d %H:%M:%S"),
                                                             str(student['Lớp khóa học']).strip(),
                                                             course_id)
                        if isStudentCreated is True:
                            new_students.append(student)
                        else:
                            current_student = Student.searchStudentRecord(str(student['Mã SV']).strip())[0]
                            Student.updateRecord(current_student['user']['user_id'], current_student['student_id'],
                                                 str(student['Mã SV']).strip(), str(student['Mã SV']).strip(),
                                                 str(student['Họ và tên']).strip(),
                                                 str(student['Mã SV']) + '@vnu.edu.vn', datetime.strptime(student['Ngày sinh'], '%d/%m/%Y').strftime("%Y-%m-%d %H:%M:%S"),
                                                 str(student['Lớp khóa học']).strip(), course_id,
                                                 datetime.now(pytz.timezone('Asia/Ho_Chi_Minh')).strftime(
                                                     "%Y-%m-%d %H:%M:%S"), 1,
                                                 0)

                            update_students.append(student)
                    else:
                        error_students.append(student)

            return jsonify({'status': 'success',
                            'new_students': new_students,
                            'error_students': error_students,
                            'updated_students': update_students,
                            }), 200
        else:
            return jsonify(
                {'status': 'bad-request', 'error_message': excel_file + ' không đúng định dạng xlsx hoặc xls'}), 400
    except Exception as e:
        return jsonify({'status': 'bad-request', 'error_message': e.__str__()}), 400


@student.route('/records', methods=['GET'])
@token_required
def get_records(e):
    try:
        page_index = request.args.get('page_index')
        per_page = request.args.get('per_page')
        sort_field = request.args.get('sort_field')
        sort_order = request.args.get('sort_order')
        record = Student.getRecord(page_index, per_page, sort_field, sort_order)

        return jsonify({
            'status': 'success',
            'records': record[0],
            'page_number': record[1].page_number,
            'page_size': record[1].page_size,
            'num_pages': record[1].num_pages,
            'total_results': record[1].total_results
        }), 200
    except Exception as e:
        return jsonify({'status': 'bad-request', 'error_message': e.__str__()}), 400


@student.route('/records-by-course', methods=['GET'])
@token_required
def get_records_by_course(e):
    try:
        course_id = request.args.get('course_id')
        page_index = request.args.get('page_index')
        per_page = request.args.get('per_page')
        sort_field = request.args.get('sort_field')
        sort_order = request.args.get('sort_order')
        record = Student.getRecordByCourse(course_id, page_index, per_page, sort_field, sort_order)

        return jsonify({
            'status': 'success',
            'records': record[0],
            'page_number': record[1].page_number,
            'page_size': record[1].page_size,
            'num_pages': record[1].num_pages,
            'total_results': record[1].total_results
        }), 200
    except Exception as e:
        return jsonify({'status': 'bad-request', 'error_message': e.__str__()}), 400


@student.route('/search', methods=['GET'])
@token_required
def search_record(e):
    try:
        searchCode = request.args.get('searchCode')
        searchRecord = Student.searchStudentRecord(str(searchCode))

        return jsonify({
            'status': 'success',
            'search_results': searchRecord,
        }), 200
    except Exception as e:
        return jsonify({'status': 'bad-request', 'error_message': e.__str__()}), 400


@student.route('/search-from-course', methods=['GET'])
@token_required
def search_record_from_course(e):
    try:
        course_id = request.args.get('course_id')
        searchCode = request.args.get('searchCode')
        searchRecord = Student.searchStudentRecordFromCourse(course_id, str(searchCode), 'in_course')

        return jsonify({
            'status': 'success',
            'search_results': searchRecord,
        }), 200
    except Exception as e:
        return jsonify({'status': 'bad-request', 'error_message': e.__str__()}), 400


@student.route('/search-from-course-existence', methods=['GET'])
@token_required
def search_from_course_existence(e):
    try:
        searchCode = request.args.get('searchCode')
        searchRecord = Student.searchStudentFromCourseExistence(str(searchCode))

        return jsonify({
            'status': 'success',
            'search_result': searchRecord,
        }), 200
    except Exception as e:
        return jsonify({'status': 'bad-request', 'error_message': e.__str__()}), 400


@student.route('/update-record', methods=['PUT'])
@token_required
def update_record(e):
    try:
        new_update = request.get_json()
        user_id = new_update.get('user_id')
        student_id = new_update.get('student_id')
        code = new_update.get('update_code')
        username = new_update.get('update_username')
        name = new_update.get('update_name')
        email = new_update.get('update_email')
        dob = new_update.get('update_dob')
        class_course = new_update.get('update_class_course')
        course_id = new_update.get('update_course_id')
        updated_at = datetime.now(pytz.timezone('Asia/Ho_Chi_Minh')).strftime("%Y-%m-%d %H:%M:%S")
        actived = new_update.get('update_actived')
        is_lock = new_update.get('update_is_lock')

        isUpdated = Student.updateRecord(int(user_id), int(student_id), str(code).strip(), str(username).strip(),
                                         str(name).strip(), str(email).strip(), dob, str(class_course).strip(),
                                         course_id, updated_at,
                                         actived, is_lock)
        if isUpdated is True:
            return jsonify({'status': 'success'}), 200
        else:
            return jsonify({'status': 'already-exist'}), 202
    except Exception as e:
        return jsonify({'status': 'bad-request', 'error_message': e.__str__()}), 400


@student.route('/delete-record', methods=['DELETE'])
@token_required
def delete(e):
    try:
        delStudent = request.get_json()
        user_id = delStudent.get('delUserID')
        Student.deleteRecord(user_id)

        return jsonify({'status': 'success'}), 200
    except Exception as e:
        return jsonify({'status': 'bad-request', 'error_message': e.__str__()}), 400


@student.route('/create-student-course-record', methods=['POST'])
@token_required
def create_student_course(e):
    try:
        new_student_course = request.get_json()
        student_id = new_student_course.get('student_id')
        course_id = new_student_course.get('course_id')
        Student.createRecordByCourse(student_id, course_id)

        return jsonify({'status': 'success'}), 200
    except Exception as e:
        return jsonify({'status': 'bad-request', 'error_message': e.__str__()}), 400


@student.route('/update-student-course-record', methods=['PUT'])
@token_required
def update_student_course(e):
    try:
        new_student_course = request.get_json()
        student_id = new_student_course.get('student_id')
        current_course_id = new_student_course.get('current_course_id')
        update_course_id = new_student_course.get('update_course_id')
        Student.updateRecordByCourse(student_id, current_course_id, update_course_id)

        return jsonify({'status': 'success'}), 200
    except Exception as e:
        return jsonify({'status': 'bad-request', 'error_message': e.__str__()}), 400


@student.route('/delete-student-course-record', methods=['DELETE'])
@token_required
def delete_student_course(e):
    try:
        delStudent = request.get_json()
        course_id = delStudent.get('course_id')
        student_id = delStudent.get('student_id')
        Student.deleteRecordByCourse(student_id, course_id)

        return jsonify({'status': 'success'}), 200
    except Exception as e:
        return jsonify({'status': 'bad-request', 'error_message': e.__str__()}), 400
